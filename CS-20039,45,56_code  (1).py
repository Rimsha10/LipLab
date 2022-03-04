import pandas as pd
import datetime
import openpyxl
from abc import ABC, abstractmethod
import xlsxwriter

date_time_day = datetime.datetime.now()


class User(ABC):

    @abstractmethod
    def signin(self):
        pass

    @abstractmethod
    def signup(self):
        pass


class Customer(User):
    count=0
    @staticmethod
    def read_customer_account():
        sheet1 = pd.read_excel('Customer.xlsx', sheet_name=0)
        list_of_details_in_dict = sheet1.to_dict('list')
        return list_of_details_in_dict


    @staticmethod
    def account_check(email, password, allinfo_dict):
        flag = 0
        pass_satisfied = 0
        info_2 = Admin.read_admin_account()
        if email in info_2['E-mail']:
            index_email_admin = info_2['E-mail'].index(email)
            if password == info_2['Password'][index_email_admin]:
                print('Admin ID cannot be used to sign in as customer\n')
                return
        if email in allinfo_dict['E-mail']:
            flag += 1
            index_email = allinfo_dict['E-mail'].index(email)
            if password == allinfo_dict['Password'][index_email]:
                pass_satisfied += 1
                print(f"\nLogged in Successfully!\nWelcome Back {allinfo_dict['First Name'][index_email]}!")
                return index_email
        if flag == 0 and pass_satisfied == 0:
            print('\nThis ID doesnot exist!')
        elif flag == 0:
            print('Incorrect ID\nPlease try again!\n')
        elif pass_satisfied == 0:
            print('Incorrect Password\nPlease try again!\n')


    def signin(self):
        Customer.count=0
        print('\nSIGN IN')
        your_email = input('\nEnter your Email Address: ')
        password = input('Enter your Password: ')
        self.info_1 = Customer.read_customer_account()
        customer_serialno = Customer.account_check(your_email, password, self.info_1)
        if type(customer_serialno) == int:
            df = pd.DataFrame(self.info_1)
            active_customer_dataframe = df.iloc[customer_serialno]
            active_user= active_customer_dataframe.to_dict()
            Customer.count += 1
            return active_user
        else:
            return
    def signup(self):
       Customer.count = 0
       customer_accounts_sheet = Customer.read_customer_account()
       info_2 = Admin.read_admin_account()
       print('SIGN UP\n')
       f_name = input('Enter First Name: ')
       l_name = input('Enter Last Name: ')
       email_address = input('Enter Email Address: ')
       password = input('Enter Password: ')
       if email_address in info_2['E-mail']:
            index_email_admin = info_2['E-mail'].index(email_address)
            if password == info_2['Password'][index_email_admin]:
                print('Admin ID cannot be used to sign up as customer\n')
                return
       if email_address in customer_accounts_sheet['E-mail']:
            print('This email is already registered.\nPlease sign in with your ID')
            return
       else:
        mobile_number = input('Enter your Mobile Number:')
        address = input('Enter your Residential Address:')
        payment_type = int(input('1.VisaCard\n2.MasterCard\n3.None\nEnter your card type:'))
        if payment_type ==1 or payment_type==2:
            cardno = input('Enter your card number: ')
            if payment_type == 1:
                payment_type = 'VisaCard'
            if payment_type == 2:
                payment_type = 'MasterCard'
        if payment_type==3:
                payment_type = 'None'
                cardno='None'
        customer_accounts_sheet = Customer.read_customer_account()
        customer_serialno = customer_accounts_sheet['S.No.'][-1] + 1
        customer_accounts_sheet['S.No.'].append(customer_serialno)
        customer_accounts_sheet['First Name'].append(f_name)
        customer_accounts_sheet['Last Name'].append(l_name)
        customer_accounts_sheet['E-mail'].append(email_address)
        customer_accounts_sheet['Password'].append(password)
        customer_accounts_sheet['Mobile Number'].append(mobile_number)
        customer_accounts_sheet['Address'].append(address)
        customer_accounts_sheet['Payment Type'].append(payment_type)
        customer_accounts_sheet['Card Number'].append(cardno)
        customer_accounts_DataFrame = pd.DataFrame(customer_accounts_sheet)
        index=customer_accounts_sheet['S.No.'].index(customer_serialno)
        writer = pd.ExcelWriter("Customer.xlsx", engine='xlsxwriter', mode='w')
        customer_accounts_DataFrame.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        active_customer_dataframe = customer_accounts_DataFrame.iloc[index]
        active_user = active_customer_dataframe.to_dict()
        Customer.count += 1
        return active_user



class Admin(User):
    @staticmethod
    def read_admin_account():
        sheet2 = pd.read_excel('Admin.xlsx', sheet_name=0)
        list_of_details_in_dict = sheet2.to_dict('list')
        return list_of_details_in_dict

    @staticmethod
    def account_check(email, password, allinfo_dict):
        flag = 0
        pass_satisfied = 0
        if email in allinfo_dict['E-mail']:
            flag += 1
            index_email = allinfo_dict['E-mail'].index(email)
            if password == allinfo_dict['Password'][index_email]:
                pass_satisfied += 1
                print(f"\nLogged in Successfully!\nWelcome Back {allinfo_dict['First Name'][index_email]}!\n")
                return index_email
        if flag == 0 and pass_satisfied == 0:
            print('\nThis ID doesnot exist!\nPlease Sign Up\n')
        elif flag == 0:
            print('Incorrect ID')
        elif pass_satisfied == 0:
            print('Incorrect Password')
    def user_menu(self):
      Admin.user=0
      while True:
       print('Welcome! visit the application\n1.As Customer\n2.As Admin\n')
       user_choice = int(input('Enter an option: '))
       if user_choice==1 or user_choice==2:
         print('\n1.Sign In\n2.Sign Up\n3.Exit')
         self.register_option = int(input('Enter an option to continue: '))
         if (user_choice)==1 and (self.register_option) == 1:
            self.active_user=Customer.signin(self)
            if type(self.active_user)!=dict:
                continue
            else:
                break
         if user_choice == 1 and self.register_option == 2:
            self.active_user=Customer.signup(self)
            if type(self.active_user) != dict:
                continue
            else:
                break
         if (user_choice) ==2 and (self.register_option) == 2:
            self.active_user=Admin.signup(self)
            if type(self.active_user) != dict:
                continue
            else:
                break
         if user_choice == 2 and self.register_option == 1:
            self.active_user = Admin.signin(self)
            if type(self.active_user) !=dict:
                continue
            else:
                break
         if self.register_option == 3:
            print('\nThank you for visting!\n')
            break
         else:
            print('Please enter a valid option')
       else:
          print('Please enter a valid option')

    def signin(self):
        print('\nSIGN IN')
        your_email = input('\nEnter your Email Address: ')
        password = input('Enter your Password: ')
        info_2 = Admin.read_admin_account()
        admin_serialno = Admin.account_check(your_email, password,info_2)
        if type(admin_serialno) == int:
            df = pd.DataFrame(info_2)
            active_admin_dataframe = df.iloc[admin_serialno]
            active_user = active_admin_dataframe.to_dict()
            return active_user
        else:
            return
    def signup(self):
      print('\nSIGN UP')
      customer_accounts_sheet=Customer.read_admin_account()
      f_name = input('Enter First Name: ')
      l_name = input('Enter Last Name: ')
      email_address = input('Enter Email Address: ')
      if email_address in customer_accounts_sheet['E-mail']:
            print('This email is already registered.\nPlease sign in with your ID')
            return
      else:
        password = input('Enter Password: ')
        admin_account_details = Admin.read_admin_account()
        admin_serialno = admin_account_details['S.No.'][-1] + 1
        admin_account_details['S.No.'].append(admin_serialno)
        admin_account_details['First Name'].append(f_name)
        admin_account_details['Last Name'].append(l_name)
        admin_account_details['E-mail'].append(email_address)
        admin_account_details['Password'].append(password)
        admin_account_dataframe = pd.DataFrame(admin_account_details)
        writer = pd.ExcelWriter("Admin.xlsx", engine='xlsxwriter', mode='w')
        admin_account_dataframe.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        index=admin_account_details['S.No.'].index(admin_serialno)
        active_admin_dataframe = admin_account_dataframe.iloc[index]
        active_user = active_admin_dataframe.to_dict()
        return active_user
    def get_user(self):
        get_user=self.active_user
        return get_user


    def add_product(self):
        WB = openpyxl.load_workbook('Product.xlsx')
        WS = WB['Product']
        while True:
            Product_ID = int(input('Set the ID of new product = '))
            Name = input('Set name of product = ')
            Price = int(input('Set price of product = '))
            Quantity_Available = int(input('Add quantity to stock = '))
            WS.append([Product_ID, Name, Price, Quantity_Available])
            WB.save(filename='Product.xlsx')
            Exit = input('To exit, press n/N\n To continue adding press y/Y = ')
            if Exit == 'n' or Exit == 'N':
                break
        print('\nITEM ADDED SUCCESSFULLY')

    def remove_product(self):
        file = pd.read_excel('Product.xlsx', index_col=[0])
        while True:
            ID = int(input('Enter ID of a product to be removed = '))
            file = file.drop(ID)
            Exit = input('To exit, press n/N\n To continue removing press y/Y = ')
            if Exit == 'n' or Exit == 'N':
                break
        file.to_excel('Product.xlsx')
        print(file)
        print('\nITEM REMOVED SUCCESSFULLY')

    def change_privacy_policy(self):
        while True:
            choice = int(input(
                '\n1.Add new policy\n2.Completely change the policy\n3.Read Policy\n4.Go to Main Menu\nSelect option number = '))
            if choice == 1:
                words = input('Write here \n__')
                with open('Privacy policy.txt', 'a+') as f:
                    f.write('\n')
                    f.write(words)
                    print('\nPOLICY CHANGED SUCCESSFULLY')
                print()
            if choice == 2:
                words = input('Write here \n__')
                with open('Privacy policy.txt', 'w+') as f:
                    f.write(words)
                    print('\nPOLICY CHANGED SUCCESSFULLY')
            if choice == 3:
                with open('Privacy policy.txt', 'r') as f:
                    for words in f:
                        print(words)
                print()
            if choice == 4:
                break


class Menu(ABC):
    @abstractmethod
    def logout(self):
        pass

    @abstractmethod
    def editAccountDetails(self):
        pass

    @abstractmethod
    def deleteAccount(self):
        pass

    @abstractmethod
    def display(self):
        pass

    def show_all_products(self):
        read_products = pd.read_excel('Product.xlsx', sheet_name=0,index_col=[0])
        print(read_products)


class CustomerMenu(Menu):
    '''  1.About Us
    2.Products->  1.Display My Profile 2.Delete Account 3.Change Account Details 4.Show Shopping History 5.Logout 6.Go to Main Menu
    3.Profile->1.Display My Profile 2.Delete Account 3.Change Account Details 4.Logout 5.Go to Main Menu
    4.My Shopping Cart
    5.Lip Lab
    6.FAQs
    7.Privacy Policy
    8.Exit'''

    def __init__(self, customer,get_active_customer):
        self.customer=customer
        self.active_customer = get_active_customer


    def DisplayMenu(self):
        objs = ShoppingCart(self.customer,self.active_customer)
        count=0
        while True:
            if count>0:
              break
            print('\nMAIN MENU\n1.About Us\n2.Products\n3.Profile\n4.My Shopping Cart\n5.Lip Lab\n6.FAQ\'S\n7.Privacy Policy\n8.Exit\n')
            menuoption = int(input('Enter an option: '))
            if menuoption == 1:
                self.about()
            if menuoption == 2:
                while True:
                    print('')
                    print('                               PRODUCTS MENU\n')
                    self.show_all_products()
                    add_to_cart_option = input('\nDo you want to add a product to your cart? [Y/y or N/n]  __')
                    print('')
                    if add_to_cart_option == 'Y' or add_to_cart_option == 'y':
                        self.add_to_cart()
                        shopping_option = int(input('Would you like to:\n1.Go to Shopping Cart options\n2.Continue Shopping\n___'))
                        if shopping_option == 1:
                            objs.cart_menu_display()
                            break
                        if shopping_option == 2:
                            continue
                    elif add_to_cart_option == 'N' or add_to_cart_option == 'n':
                        main_menu_option = input('\nDo you want to go to the Main Menu? [Y/y or N/n]  __')
                        if main_menu_option == 'Y' or main_menu_option == 'y':
                            break
                        elif main_menu_option == 'N' or main_menu_option == 'n':
                            print('\nContinue to your shopping!!\n')
                    else:
                        print('Please enter a valid option')
            if menuoption == 3:
                while True:
                    print(
                        '\n\nPROFILE\n1.Display My Profile\n2.Delete Account\n3.Change Account Details\n4.Show Shopping History\n5.Logout\n6.Go to Main Menu')
                    profile_menu = int(input('\nEnter an option :'))
                    print('')
                    if profile_menu == 1:
                        self.display()
                    if profile_menu == 2:
                        self.deleteAccount()
                        count+=1
                        break
                    if profile_menu == 3:
                        self.editAccountDetails()
                    if profile_menu == 4:
                        self.shopping_history()
                    if profile_menu == 5:
                        self.logout()
                        count+=1
                        break

                    if profile_menu == 6:
                        break
            if menuoption == 4:
                customer_cart_name = 'customer_ID_' + str(self.active_customer['S.No.'])
                customers_cart = pd.ExcelFile("Shopping Cart.xlsx")
                if customer_cart_name in customers_cart.sheet_names:
                    print('\n')
                    objs.show_cart()
                    objs.cart_menu_display()
                else:
                    print('\nYour Shopping Cart is empty!\nFill your cart with our luscious collection of lip products!!\n')
            if menuoption == 5:
                self.LipLab()
            if menuoption == 6:
                self.FAQS()
            if menuoption == 7:
                self.privacy_policy()
            if menuoption == 8:
                print('\nThank you for visting!!\nHave a nice day!\n')
                break

    @staticmethod
    def read_products_file():
        products_data = pd.read_excel('Product.xlsx', na_values="Missing", sheet_name=0)
        products_dataframe = pd.DataFrame(products_data)
        return products_dataframe

    def logout(self):
        print('\nLogged out successfully!\nThank you for visiting the LIP WORLD')

    def editAccountDetails(self):
        customer_account = Customer.read_customer_account()
        print(
            'Search By field:\n1.First Name\n2.Last Name\n3.Email Address\n4.Password\n5.Mobile Number\n6.Residential Address\n7.Payment Method\n8.Card Number')
        choice = int(input('\nEnter an option: '))
        print('')
        changes = input('Enter changes  for the required field: ')
        if choice == 1:
            self.active_customer['First Name'] = changes
        if choice == 2:
            self.active_customer['Last Name'] = changes
        if choice == 3:
            self.active_customer['E-mail'] = changes
        if choice == 4:
            self.active_customer['Password'] = changes
        if choice == 5:
            self.active_customer['Mobile Number'] = changes
        if choice == 6:
            self.active_customer['Address'] = changes
        if choice == 7:
            self.active_customer['Payment Type'] = changes
        if choice == 8:
            self.active_customer['Card Number'] = changes
        row = list(self.active_customer.values())
        workbook = openpyxl.load_workbook("Customer.xlsx")
        ws = workbook.active
        index=(customer_account['S.No.'].index(self.active_customer['S.No.']))+2
        for i in range(1, 10):
            ws.cell(row=index, column=i, value=row[i - 1])
        workbook.save("Customer.xlsx")
        print('Account details has been edited succeccfully!\n')

    def display(self):
        displayprofile = self.active_customer
        print('\nACCOUNT DETAILS\n')
        print('First Name: ', displayprofile['First Name'])
        print('Last Name: ', displayprofile['Last Name'])
        print('Email Address: ', displayprofile['E-mail'])
        print('Password: ', len(displayprofile['Password']) * '*')
        print('Mobile Number: ', displayprofile['Mobile Number'])
        print('Address: ', displayprofile['Address'])
        print('Payment Type: ', displayprofile['Payment Type'])
        print('Card Number: ', displayprofile['Card Number'], '\n')

    def about(self):
        f = open('About.txt')
        content = f.read()
        print(content)
        f.close()

    def FAQS(self):
        f = open('FAQS.txt', 'r')
        content = f.read()
        print('*'*150)
        print(content)
        f.close()
        print('*' * 150)

    def LipLab(self):
        f = open('LipLab.txt', 'r')
        content = f.read()
        print('*' * 105)
        print(content)
        f.close()
        print('*' * 105)

    def privacy_policy(self):
        with open('Privacy policy.txt', 'r') as f:  # closes file automatically
            for words in f:
                print(words)

    def add_to_cart(self):
        products_dataframe = CustomerMenu.read_products_file()
        dict_of_products = products_dataframe.to_dict('list')
        product_id = int(input('\nEnter product ID to add to cart: '))

        if product_id in dict_of_products['Product_ID']:
            index = dict_of_products['Product_ID'].index(product_id)
            df = pd.DataFrame(dict_of_products)
            product_dataframe = df.iloc[index:index + 1]
            self.active_product = product_dataframe.to_dict('list')

            if self.active_product['Quantity Available'][0] > 0:
                while True:
                    try:
                        quantity = int(input('Quantity : '))
                        if self.active_product['Quantity Available'][0] < quantity  or quantity<=0:
                            raise ValueError('\nYou can add product quantity only within "Available Quantity"\n ')
                        break
                    except ValueError as e1:
                        print(e1)

                self.active_product['Quantity'] = self.active_product.pop('Quantity Available')
                self.active_product['Quantity'][0] = quantity

                customer_cart_name = 'customer_ID_' + str(self.active_customer['S.No.'])
                customer_cart_file = pd.read_excel("Shopping Cart.xlsx", sheet_name=None)
                customers_cart = pd.ExcelFile("Shopping Cart.xlsx")

                if customer_cart_name in customers_cart.sheet_names:
                    customer_cart_dict = customer_cart_file[customer_cart_name].to_dict('list')
                    WB = openpyxl.load_workbook("Shopping Cart.xlsx")
                    if product_id in customer_cart_dict['Product_ID']:
                        id_index = customer_cart_dict['Product_ID'].index(product_id)
                        customer_cart_dict['Quantity'][id_index] = customer_cart_dict['Quantity'][id_index] + quantity
                        worksheet = WB.active
                        worksheet.cell(row=id_index + 2, column=4, value=customer_cart_dict['Quantity'][id_index])
                        WB.save(filename="Shopping Cart.xlsx")
                    else:
                        sheet = WB[customer_cart_name]
                        sheet.append([self.active_product['Product_ID'][0], self.active_product['Name'][0],
                                      self.active_product['Price'][0], self.active_product['Quantity'][0]])
                        WB.save(filename="Shopping Cart.xlsx")
                else:
                    header = list(self.active_product.keys())
                    row = list(self.active_product.values())
                    workbook = openpyxl.load_workbook("Shopping Cart.xlsx")
                    worksheet1 = workbook.create_sheet()
                    worksheet1.title = customer_cart_name
                    for i in range(1, 5):
                        worksheet1.cell(row=1, column=i, value=header[i - 1])
                        worksheet1.cell(row=2, column=i, value=row[i - 1][0])
                    workbook.save(filename="Shopping Cart.xlsx")
                print('\nProduct has been added to your shopping cart successfully!\n')

            else:
                print('\nSorry the product is of stock!\n')
        else:
            print('\nYou have entered an invalid Product ID\n')

    def shopping_history(self):
        customer_serial_no = self.active_customer['S.No.']
        customer_history_sheet_name = 'customer_ID_'  + str(customer_serial_no)
        list_of_customers = pd.ExcelFile('Shopping History.xlsx')
        if customer_history_sheet_name in list_of_customers.sheet_names:
            shopping_history_file = pd.read_excel('Shopping History.xlsx', na_values="Missing",
                                                  sheet_name=None)
            shopping_history_dict = shopping_history_file[customer_history_sheet_name].to_dict('records')
            shopping_history_dataframe = pd.DataFrame(data=shopping_history_dict)
            print(shopping_history_dataframe.to_string(index=False))
        else:
            print('\nThere is no shopping history!\n')

    def deleteAccount(self):
        delete_customer_account = Customer.read_customer_account()
        index=delete_customer_account['S.No.'].index(self.active_customer['S.No.'])
        for keys in delete_customer_account:
            delete_customer_account[keys].pop(index)
        updated_df = pd.DataFrame(delete_customer_account)
        writer = pd.ExcelWriter("Customer.xlsx", engine='xlsxwriter', mode='w')
        updated_df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        customer_cart_name = 'customer_ID_' + str(self.active_customer['S.No.'])
        customers_cart = pd.ExcelFile("Shopping Cart.xlsx")
        if customer_cart_name in customers_cart.sheet_names:
          cart_workbook = openpyxl.load_workbook('Shopping Cart.xlsx')
          cart_workbook.remove(cart_workbook['customer_ID_' + str(self.active_customer['S.No.'])])
          cart_workbook.save(filename="Shopping Cart.xlsx")
        customers_history = pd.ExcelFile("Shopping History.xlsx")
        if customer_cart_name in customers_history.sheet_names:
          history_workbook = openpyxl.load_workbook('Shopping History.xlsx')
          history_workbook.remove(history_workbook['customer_ID_' + str(self.active_customer['S.No.'])])
          history_workbook.save(filename="Shopping History.xlsx")

        print('Your account has been deleted!\nPlease sign up or sign in with another ID\n')




class AdminMenu(Menu):
    '''1.Profile->1.Display My Profile, 2.Delete Account, 3.Change Account Details, 4.Logout, 5.Go to Main Menu
     2.Products ->1.Add new product to stock, 2.Remove product from the stock, 3.Exit
     3.FAQ'S
     4.About
     5.Users->1.Show all customers, 2.Show all admins, 3.Exit
     6.Privacy policy->1.Add new policy, 2.Completely change the policy, 3.Read Policy, 4.Go to Main Menu
     7.Exit'''

    def __init__(self, get_admin):
        self.active_admin=get_admin


    def DisplayMenu(self):
        count=0
        while True:
            if count>0:
                break
            print('\nMENU\n1.About Us\n2.Products\n3.Profile\n4.FAQ\'S\n5.Users\n6.Privacy policy\n7.Exit')
            menuoption = int(input('\nEnter an option: '))
            if menuoption == 1:
                self.about()
            if menuoption == 2:
                while True:
                    self.show_all_products()
                    admin_option_1 = input('\nMake changes in the stocked products? [Y/y or N/n]  __')
                    if admin_option_1 == 'y' or admin_option_1 == 'Y':
                        admin_option_2 = int(input(
                            '\n1.Add new product to stock\n2.Remove product from the stock\n3.Exit\nSelect option number = '))
                        if admin_option_2 == 1:
                            self.add_product()
                        if admin_option_2 == 2:
                            self.remove_product()
                        if admin_option_2 == 3:
                            break
                    if admin_option_1 == 'n' or admin_option_1 == 'N':
                        break
            if menuoption == 3:
                while True:
                    print(
                        '\n\nPROFILE:\n1.Display My Profile\n2.Delete Account\n3.Change Account Details\n4.Logout\n5.Go to Main Menu\n')
                    profile_menu = int(input('Enter an option :'))
                    if profile_menu == 1:
                        self.display()
                    if profile_menu == 2:
                        self.deleteAccount()
                        count+=1
                        break
                    if profile_menu == 3:
                        self.editAccountDetails()
                    if profile_menu == 4:
                        self.logout()
                        count+=1
                        break
                    if profile_menu == 5:
                        break
            if menuoption == 4:
                self.FAQS()
            if menuoption == 5:
                while True:
                    admin_option_3 = int(
                        input('\n1.Show all customers\n2.Show all admins\n3.Exit\nSelect option number = '))
                    if admin_option_3 == 1:
                        self.show_all_customers()
                    if admin_option_3 == 2:
                        self.show_all_admins()
                    if admin_option_3 == 3:
                        break
            if menuoption == 6:
                Admin.change_privacy_policy(self)
            if menuoption == 7:
                print('Thank you for visting the LIP WORLD!!\nHave a nice day!')
                break

    def about(self):
        f = open('About.txt')
        content = f.read()
        print(content)
        f.close()

    def FAQS(self):
        f = open('FAQS.txt')
        content = f.read()
        print(content)
        f.close()

    def logout(self):
        print('\nLogged out successfully!\nThank you Admin')

    def editAccountDetails(self):
        admin_account=Admin.read_admin_account()
        choice = int(input('1.First Name\n2.Last Name\n3.Email\n4.Password\nEnter an option: '))
        changes = input('Enter changes to the required field = ')
        if choice == 1:
            self.active_admin['First Name'] = changes
        if choice == 2:
            self.active_admin['Last Name'] = changes
        if choice == 3:
            self.active_admin['Email'] = changes
        if choice == 4:
            self.active_admin['Password'] = changes
        row = list(self.active_admin.values())
        workbook = openpyxl.load_workbook("Admin.xlsx")
        ws = workbook.active
        index = (admin_account['S.No.'].index(self.active_admin['S.No.'])) + 2
        print(index)
        for i in range(1, 6):
            ws.cell(row= index, column=i, value=row[i - 1])
        workbook.save("Admin.xlsx")

    def display(self):
        displayprofile = self.active_admin
        print('ACCOUNT DETAILS')
        print('First Name:     ', displayprofile['First Name'])
        print('Last Name:      ', displayprofile['Last Name'])
        print('Email Address:  ', displayprofile['E-mail'])
        print('Password:       ', len(displayprofile['Password']) * '*')

    def deleteAccount(self):
        delete_admin_account = Admin.read_admin_account()
        index=delete_admin_account['S.No.'].index(self.active_admin['S.No.'])
        for keys in delete_admin_account:
            delete_admin_account[keys].pop(index)
        updated_df = pd.DataFrame(delete_admin_account)
        writer = pd.ExcelWriter("Admin.xlsx", engine='xlsxwriter', mode='w')
        updated_df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        print('Your account has been deleted!\nPlease visit the application with another ID\n')

    def show_all_customers(self):
        read_customers = pd.read_excel('Customer.xlsx', usecols='A:D')
        print(read_customers)

    def show_all_admins(self):
        read_admins = pd.read_excel('Admin.xlsx', usecols='A:D')
        print(read_admins)


class ShoppingCart:
    def __init__(self, customer,get_user):
        self.obj_b=Bill(customer,get_user)
        self.active_customer_2 = get_user

    @staticmethod
    def read_cart_file(filename):
        read_cart_file = pd.read_excel("Shopping Cart.xlsx",na_values="Missing", sheet_name=None,)
        cart = read_cart_file['customer_ID_' + filename]
        cart_dataframe = pd.DataFrame(cart)
        return cart_dataframe

    def cart_menu_display(self):

        count=0
        while True:
          if count==1:
            break
          else:
            print(
                '\nSHOPPING CART MENU\n1.Show My Cart\n2.Edit Shopping Cart\n3.Place Order\n4.Go to Previous Menu\n')
            input_choice = int(input('Enter an option number: '))
            if input_choice == 1:
                self.show_cart()
            if input_choice == 2:
               while True:
                 print('MY SHOPPING CART')
                 if self.show_cart()==False:
                    count+=1
                    break
                 else:
                    cart_dataframe = ShoppingCart.read_cart_file(str(self.active_customer_2['S.No.']))
                    cart_dict = cart_dataframe.to_dict('list')
                    self.id = int(input('Enter a Product ID:  '))
                    if self.id in cart_dict['Product_ID']:
                        self.id_index = cart_dict['Product_ID'].index(self.id)
                        print(
                            '\nEDIT SHOPPING CART \n\n1.Remove Product\n2.Add Quantity\n3.Reduce Quantity\n4.Go to Shopping Cart Menu')
                        option = int(input('Enter an option number to edit your cart: '))
                        if option == 1:
                            self.delete_product()
                        if option == 2:
                            self.add_quantity()
                        if option == 3:
                            self.remove_quantity()
                        if option == 4:
                            break
                    else:
                        print('This product is not in your cart\nPlease enter a valid Product ID\n')
                        continue
            if input_choice == 4:
                break
            if input_choice == 3:
                self.obj_b.bill_reciept()
                break

    def show_cart(self):
        dataframe=ShoppingCart.read_cart_file(str(self.active_customer_2['S.No.']))

        dict=dataframe.to_dict('list')
        if dict['Product_ID']==[]:
            print('no items in the cart\nFill your cart with luscious LIP products')
            return False
        else:
          print(dataframe.to_string(index=False))
          print(' ')

    def delete_product(self):
        cart_dataframe = ShoppingCart.read_cart_file(str(self.active_customer_2['S.No.']))
        cart_dict = cart_dataframe.to_dict('list')
        ID_index = cart_dict['Product_ID'].index(self.id)
        book = openpyxl.load_workbook("Shopping Cart.xlsx")
        sheet = book['customer_ID_' + str(self.active_customer_2['S.No.'])]
        sheet.delete_rows(ID_index + 2)
        book.save(filename="Shopping Cart.xlsx")
        print('Product has been removed successfully!\n')

    def delete_cart(self):

        workbook = openpyxl.load_workbook("Shopping Cart.xlsx")
        workbook.remove(workbook['customer_ID_' + str(self.active_customer_2['S.No.'])])
        workbook.save("Shopping Cart.xlsx")
        print('\nYour Shopping Cart is empty now!\n')

    def remove_quantity(self):
        cart_dataframe =ShoppingCart.read_cart_file(str(self.active_customer_2['S.No.']))

        cart_dict = cart_dataframe.to_dict('list')
        workbook = openpyxl.load_workbook("Shopping Cart.xlsx")
        worksheet = workbook.active
        sheet = workbook['customer_ID_'+str(self.active_customer_2['S.No.'])]
        get_cell = sheet.cell(row=self.id_index + 2, column=4).value
        while True:
            try:
                quantity = int(input('Enter quantity: '))
                if cart_dict['Quantity'][self.id_index] < quantity:
                    raise ValueError('\nYou can reduce product quantity to zero only\nPlease enter a valid quantity\n ')
                break
            except ValueError as e1:
                print(e1, '\n')
        if quantity == cart_dict['Quantity'][self.id_index]:
            self.delete_product()
        else:
            sheet.cell(row=self.id_index + 2, column=4, value=get_cell - quantity)
            workbook.save(filename="Shopping Cart.xlsx")
            print('\nQuantity of product has been reduced\n')

        print('')

    def add_quantity(self):
        active = CustomerMenu.read_products_file()
        active_products_dict = active.to_dict('list')
        workbook = openpyxl.load_workbook("Shopping Cart.xlsx")
        worksheet = workbook.active
        sheet = workbook['customer_ID_' + str(self.active_customer_2['S.No.'])]
        get_cell = sheet.cell(row=self.id_index + 2, column=4).value
        while True:
            try:
                quantity = int(input('Enter quantity: '))
                if active_products_dict['Quantity Available'][active_products_dict['Product_ID'].index(self.id)] < quantity + get_cell:
                    raise ValueError('You can add product quantity only within "Available Quantity"\n ')
                break
            except ValueError as e1:
                print(e1,'\n')
        sheet.cell(row=self.id_index + 2, column=4, value=get_cell + quantity)
        workbook.save(filename="Shopping Cart.xlsx")
        print('\nQuantity of product has added successfully\n')


class Bill:
    """Bill class to confirm the order and get a reciept of shopping"""

    def __init__(self, customer,get_user):
        self.customer = customer
        self.active_customer_3 = get_user

    def bill_reciept(self):
        self.cart_workbook = openpyxl.load_workbook('Shopping Cart.xlsx')
        self.customer_number = 'customer_ID_' + str(self.active_customer_3['S.No.'])
        if self.customer_number in self.cart_workbook.sheetnames:            # checking if he/she has added products or not
            read_order = pd.read_excel('Shopping Cart.xlsx', sheet_name=self.customer_number,index_col=[0])
            read_order['Total Price'] = read_order.apply(lambda row: (row['Quantity']*row['Price']),axis=1)
            self.bill = read_order['Total Price'].sum(axis=0, skipna=False)    # adding the prices of all products this column
            print('Your order\n', read_order, '\n')
            print('Is confirmed at', date_time_day.strftime("%X"), 'On', date_time_day.strftime("%A"),
                  date_time_day.strftime("%x"), '\n')
            print('Your Total bill is = $', self.bill)
            if self.active_customer_3['Payment Type']=='None':
                paymentoption=int(input('Please specifiy Payment method:\n1.Card\n(we accept MasterCard/VisaCard only)\n2.COD\n __'))
                if paymentoption==1:
                    card_type =int(input('1.VisaCard\n2.MasterCard\nEnter your card type:_ '))
                    if card_type ==1 or card_type==2:
                        cardno = input('Enter your card number: ')
            else:
                print(f"\nYour Payment has been deducted from your {self.active_customer_3['Payment Type']} of no. {self.active_customer_3['Card Number']}\n")
            print('THANK YOU FOR SHOPPING WITH US')
            self.update_quantity()
            self.update_history()
            self.delete_sheet()

        else:
            print('Your shopping cart is empty!!\nFill it with our luscious Lip makeup products now!!')
            CustomerMenu.show_all_products(self)

    def delete_sheet(self):
        #if obj == None:
            #obj = self
        cart_workbook = openpyxl.load_workbook('Shopping Cart.xlsx')
        cart_workbook.remove(cart_workbook['customer_ID_' + str(self.active_customer_3['S.No.'])])
        #cart_workbook.remove(cart_workbook['customer_ID_'+self.ind])
        cart_workbook.save(filename="Shopping Cart.xlsx")
    def update_quantity(self):
        #if obj == None:
            #obj = self
            read_order = pd.read_excel('Shopping Cart.xlsx', sheet_name=self.customer_number)
            cart_dict = read_order.to_dict('list')
            workbook = openpyxl.load_workbook("Product.xlsx")
            worksheet = workbook.active
            products_names = cart_dict['Name']
            for k in range(0,len(products_names)):
              for i in range(0, worksheet.max_row):
                get_cell = worksheet.cell(row=i + 2, column=2).value
                if products_names[k] == get_cell:
                    updated_quantity = worksheet.cell(row=i + 2, column=4).value - cart_dict['Quantity'][k]
                    worksheet.cell(row=i + 2, column=4, value=updated_quantity)
                    workbook.save(filename="Product.xlsx")
    def update_history(self):
        history = openpyxl.load_workbook('Shopping History.xlsx')
        sheets = history.sheetnames
        history_work_book = pd.read_excel('Shopping Cart.xlsx', sheet_name=self.customer_number)
        histor_book_dict = history_work_book.to_dict(
            orient='list')                                       # converting data into dictionary containing column headers as key
        del history_work_book['Product_ID']  # deleting id
        histor_book_dict['Day'] = str(date_time_day.strftime('%A'))
        histor_book_dict['Date'] = str(date_time_day.strftime('%x'))
        histor_book_dict['Time'] = str(date_time_day.strftime('%X'))
        histor_book_dict['Total Bill'] = str(self.bill)

        if self.customer_number in sheets:
            wb = openpyxl.load_workbook('Shopping History.xlsx')
            ws = wb[self.customer_number]
            row = list(histor_book_dict.values())
            for i in range(1,len(row[0])+1):
               ws.append([row[1][i-1],row[2][i-1],row[3][i-1],row[4],row[5],row[6],row[7]])
               wb.save(filename='Shopping History.xlsx')
        else:                                         # creates new work sheet for new customer
            header = list(histor_book_dict.keys())
            row = list(histor_book_dict.values())
            history_sheet = history.create_sheet()
            history_sheet.title = self.customer_number
            for i in range(1, 8):
                history_sheet.cell(row=1, column=i, value=header[i])
                history.save(filename='Shopping History.xlsx')

            for i in range(0,len(row[0])):
                   history_sheet.append([row[1][i], row[2][i], row[3][i], row[4], row[5], row[6], row[7]])
            history.save(filename='Shopping History.xlsx')



a=Admin()
a.user_menu()
if a.register_option!=3:
  c=Customer()
  if Customer.count>0:
    user = a.get_user()
    cm=CustomerMenu(c,user)
    cm.DisplayMenu()
  else:
    user = a.get_user()
    am=AdminMenu(user)
    am.DisplayMenu()
